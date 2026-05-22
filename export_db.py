"""
export_db.py - Export SQLite database tables to CSV and XLSX formats
============================================================
Run this AFTER running the system and populating the database.
Output files:
  - db_export/users.csv
  - db_export/accounts.csv
  - db_export/transfers.csv
  - db_export/audit_log.csv
  - db_export/banking_data.xlsx  (all tables in one workbook)
===========================================================================
Aaron Kalaji 10670705, CSI3344 Assignment 2
"""

import sqlite3
import csv
import os
import sys
DB_PATH = "bank.db"
EXPORT_DIR = "db_export"

def get_conn():
    if not os.path.exists(DB_PATH):
        print(f"Database '{DB_PATH}' not found, Run bdb_server.py first")
        sys.exit(1)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def export_table_to_csv(conn, table: str, filename: str):
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = cur.fetchall()
    if not rows:
        print(f"  Table '{table}' is empty.")
        return []

    headers = [description[0] for description in cur.description]
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows([list(row) for row in rows])

    print(f"  Exported {len(rows)} rows → {filename}")
    return rows, headers

def export_all_to_xlsx(conn, xlsx_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed skipping XLSX export..")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    tables = ["users", "accounts", "sessions", "transfers", "audit_log"]
    for table in tables:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM {table}")
        rows = cur.fetchall()
        if not rows:
            continue
        headers = [d[0] for d in cur.description]
        ws = wb.create_sheet(title=table.capitalize())

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(list(row), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(header)),
                *[len(str(row[col_idx - 1])) for row in rows]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        print(f"  Sheet '{table.capitalize()}': {len(rows)} rows")
    wb.save(xlsx_path)
    print(f"  Saved workbook → {xlsx_path}")

def print_summary(conn):
    """Print a summary of the database from our banking system"""
    print("\n  Database Summary")
    cur = conn.cursor()

    cur.execute("SELECT u.username, u.full_name, a.account_id, a.balance_cents "
                "FROM users u JOIN accounts a ON u.user_id = a.user_id")
    print("\n  Accounts:")
    print(f"  {'Username':<12} {'Name':<18} {'Account':<8} {'Balance':>12}")
    print(f"  {'─'*12} {'─'*18} {'─'*8} {'─'*12}")
    for row in cur.fetchall():
        bal = f"${row['balance_cents']/100:,.2f}"
        print(f"  {row['username']:<12} {row['full_name']:<18} {row['account_id']:<8} {bal:>12}")

    cur.execute("SELECT status, COUNT(*) as cnt FROM transfers GROUP BY status")
    rows = cur.fetchall()
    if rows:
        print("\n  Transfer Status Summary:")
        for row in rows:
            print(f"  {row['status']}: {row['cnt']}")
    cur.execute("SELECT COUNT(*) FROM audit_log")
    print(f"\n  Audit log entries: {cur.fetchone()[0]}")
    print("  ──────────────────────────────────────────────")

def main():
    print("=" * 52)
    print("  Export database from banking system")
    print("=" * 52)
    os.makedirs(EXPORT_DIR, exist_ok=True)
    conn = get_conn()
    print(f"\n  Exporting to '{EXPORT_DIR}/'...")

    tables = ["users", "accounts", "sessions", "transfers", "audit_log"]
    for table in tables:
        path = os.path.join(EXPORT_DIR, f"{table}.csv")
        export_table_to_csv(conn, table, path)
    xlsx_path = os.path.join(EXPORT_DIR, "banking_data.xlsx")
    export_all_to_xlsx(conn, xlsx_path)
    print_summary(conn)
    conn.close()
    print(f"\n  Export complete. Files in '{EXPORT_DIR}/'")

if __name__ == "__main__":
    main()